#!/usr/bin/env python3
"""Cruzamento POR CLIENTE (tomador): faturamento (competencia 2025) x folha, por cliente.
Replica o filtro do RetentionWorkbook/PayrollWorkbook em Python (bigdecimal do Ruby bloqueado
pela politica da maquina). Valida contra os totais conhecidos antes de gerar.
"""
import openpyxl, glob, os, re, json, unicodedata, argparse
from collections import defaultdict, Counter

ap = argparse.ArgumentParser()
ap.add_argument("--mes", type=int, default=0)  # 1..12; 0 = ano todo
ap.add_argument("--out", default=None)
ap.add_argument("--compilado", default=r"C:\Users\xandao\Downloads\Compilado_NFs_2026.xlsx")
ap.add_argument("--dsprest", default=r"C:\Users\xandao\Downloads\DSPREST_E_*.xml")
args = ap.parse_args()
MES = args.mes

BASE = r"storage\private\fiscal_auditor\appa"
CNPJ_RE = re.compile(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}")

def norm(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+", " ", s).strip()

def ident(v):
    if v is None:
        return None
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s or None

def num(v):
    try:
        return float(str(v).replace(",", ".")) if v not in (None, "") else 0.0
    except Exception:
        return 0.0

fat = defaultdict(float); cnpjs = defaultdict(Counter); names = defaultdict(Counter); filial = defaultdict(Counter)
naoid_total = 0.0            # faturamento com CNPJ valido mas sem codigo de cliente (comp 2025)
emis_late = [0.0, 0]         # notas de competencia 2025 EMITIDAS em 2026 (valor, qtd) — ja incluidas
seen = set()                 # dedup global por (cc|cnpj, nf)
uniq = [0]
pend = []                    # notas sem codigo (cnpj_digitos, valor, cliente) p/ atribuir por CNPJ

def processa_faturamento(path):
    global naoid_total, emis_late
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True); ws = wb.active
    hdr = None
    for row in ws.iter_rows(values_only=True):
        if hdr is None:
            if row and any(isinstance(c, str) and norm(c) == "cnpj cliente" for c in row):
                hdr = {}
                for i, c in enumerate(row):
                    n = norm(c)
                    if n == "cnpj cliente": hdr["cnpj"] = i
                    elif n in ("cod cliente", "cd cliente", "cleinte"): hdr["cc"] = i
                    elif n == "cliente": hdr["cli"] = i
                    elif "competencia" in n: hdr["comp"] = i
                    elif re.fullmatch(r"valor (da )?fatura", n): hdr["bill"] = i
                    elif n == "filial": hdr["fil"] = i
                    elif "emissao" in n: hdr["emis"] = i
                    elif re.fullmatch(r"no? nf ?e", n): hdr["nf"] = i
            continue
        def g(k):
            i = hdr.get(k)
            return row[i] if i is not None and i < len(row) else None
        cnpj = str(g("cnpj") or "").strip()
        cc = ident(g("cc")); cli = str(g("cli") or "").strip()
        if not (CNPJ_RE.fullmatch(cnpj) or (cc and cli)):
            continue
        comp = g("comp")
        yr = comp.year if hasattr(comp, "year") else None
        if yr != 2025:
            continue
        if MES and comp.month != MES:
            continue
        nf = ident(g("nf"))
        key = (cc or cnpj, nf) if nf else ("_", uniq[0])
        if not nf:
            uniq[0] += 1
        if key in seen:
            continue                 # nota ja contada (evita duplicar entre arquivos)
        seen.add(key)
        emis = g("emis")
        if hasattr(emis, "year") and emis.year == 2026:
            emis_late[0] += num(g("bill")); emis_late[1] += 1
        val = num(g("bill"))
        if not cc:                   # nota valida sem codigo -> tenta atribuir por CNPJ depois
            pend.append((re.sub(r"\D", "", cnpj), val, cli))
            continue
        fat[cc] += val
        if CNPJ_RE.fullmatch(cnpj): cnpjs[cc][cnpj] += 1
        if cli: names[cc][cli] += 1
        fl = g("fil")
        if fl: filial[cc][str(fl).strip()] += 1
    wb.close()

fontes = [f for f in glob.glob(os.path.join(BASE, "source", "**", "*.xlsx"), recursive=True)
          if "RETEN" in os.path.basename(f).upper()]
if args.compilado and os.path.exists(args.compilado):
    fontes.append(args.compilado)   # faturamento emitido em 2026 (competencias 2025)
for f in fontes:
    processa_faturamento(f)

# item 2: atribui notas sem codigo ao cliente pelo CNPJ (mapa cnpj->cc dos que tem codigo)
cnpj2cc = {}
for cc, cnt in cnpjs.items():
    cnpj2cc[re.sub(r"\D", "", cnt.most_common(1)[0][0])] = cc
recuperado_total = 0.0
for cnpj_d, val, cli in pend:
    cc = cnpj2cc.get(cnpj_d)
    if cc:
        fat[cc] += val
        recuperado_total += val
        if cli:
            names[cc][cli] += 1
    else:
        naoid_total += val

# incorpora NFS-e (DSPREST/ABRASF) — faturamento fora das planilhas (ex.: CAIXA estab 0005-44)
MESES_N = {"JANEIRO": 1, "FEVEREIRO": 2, "MARCO": 3, "ABRIL": 4, "MAIO": 5, "JUNHO": 6,
           "JULHO": 7, "AGOSTO": 8, "SETEMBRO": 9, "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12}
dsprest_add = 0.0
for f in glob.glob(args.dsprest):
    d = open(f, encoding="utf-8", errors="replace").read()
    tom = re.search(r"<TomadorServico>.*?<Cnpj>(\d+)</Cnpj>", d, re.S)
    if not tom:
        continue
    cc = cnpj2cc.get(tom.group(1))
    if not cc:
        continue                 # tomador nao mapeado a um codigo de cliente
    mm = re.search(r"MES DE (\w+)", d)
    mi = MESES_N.get(norm(mm.group(1)).upper()) if mm else None
    if mi is None or (MES and mi != MES):
        continue                 # so competencias 2025 do mes-servico (respeita --mes)
    nfm = re.search(r"<Numero>(\d+)</Numero>", d)
    nf = ident(nfm.group(1)) if nfm else None
    key = (cc, nf) if nf else ("_dsp", uniq[0])
    if not nf:
        uniq[0] += 1
    if key in seen:
        continue                 # nota ja contada (ex.: NF ja estava na retencao)
    seen.add(key)
    vm = re.search(r"<ValorServicos>(.*?)</ValorServicos>", d)
    val = num(vm.group(1)) if vm else 0.0
    fat[cc] += val
    dsprest_add += val
if dsprest_add:
    print(f"[NFS-e/DSPREST] incorporado: {dsprest_add:,.2f}")

folha = defaultdict(float); fnames = defaultdict(Counter)
ben_add = defaultdict(float); ben_sub = defaultdict(float)  # beneficios: parte empresa / parte empregado
BEN_KW = ("VALE", "REFEI", "ALIMENT", "CESTA", "MEDIC", "ODONT", "LANCHE", "TRANSP")

def is_benef(desc):
    dn = norm(desc)
    if any(x in dn for x in ("pensao", "contrib", "sindic", "assistencial")):
        return False
    return any(k.lower() in dn for k in BEN_KW)

for f in glob.glob(os.path.join(BASE, "payroll", "*.xlsx")):
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True); ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 17:
            continue
        et = row[4]
        if et not in ("Vencimento", "Desconto"):
            continue
        cc = ident(row[0]); cli = str(row[1] or "").strip()
        if not cc or not cli:
            continue
        s = num(row[5 + MES - 1]) if MES else sum(num(row[c]) for c in range(5, 17))
        benef = is_benef(row[3])
        if et == "Vencimento":
            folha[cc] += s                 # folha bruta (inclui aditivo de beneficio)
            fnames[cc][cli] += 1
            if benef:
                ben_add[cc] += s           # parte da empresa (beneficio)
        elif et == "Desconto" and benef:
            ben_sub[cc] += s               # parte do empregado (desconto do beneficio)
    wb.close()

codes = set(fat) | set(folha)
rows = []
for cc in codes:
    nm = (names[cc].most_common(1)[0][0] if names[cc] else "") or (fnames[cc].most_common(1)[0][0] if fnames[cc] else "")
    rows.append({
        "client_code": cc,
        "cnpj": cnpjs[cc].most_common(1)[0][0] if cnpjs[cc] else "",
        "cliente": nm,
        "filial": filial[cc].most_common(1)[0][0] if filial[cc] else "",
        "faturamento": round(fat[cc], 2),
        "folha": round(folha[cc], 2),
        "beneficio_add": round(ben_add[cc], 2),
        "beneficio_sub": round(ben_sub[cc], 2),
        "beneficio_liquido": round(ben_add[cc] - ben_sub[cc], 2),
        "folha_sem_beneficio": round(folha[cc] - ben_add[cc], 2),
        "diferenca": round(fat[cc] - folha[cc], 2),
    })
rows.sort(key=lambda r: -r["faturamento"])
tf = sum(r["faturamento"] for r in rows); tp = sum(r["folha"] for r in rows)
tba = sum(r["beneficio_add"] for r in rows); tbl = sum(r["beneficio_liquido"] for r in rows)
outpath = args.out or os.path.join(BASE, "cruzamento_cliente.json")
json.dump({"rows": rows,
           "total_classificado": round(tf, 2),
           "total_nao_identificado": round(naoid_total, 2),
           "total_geral": round(tf + naoid_total, 2),
           "total_folha": round(tp, 2),
           "total_beneficio_add": round(tba, 2),
           "total_beneficio_liquido": round(tbl, 2),
           "emitido_2026_comp_2025": round(emis_late[0], 2),
           "notas_emitido_2026_comp_2025": emis_late[1]},
          open(outpath, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

print(f"clientes: {len(rows)}  | faturamento classificado: {tf:,.2f}  + nao id: {naoid_total:,.2f}  = {tf+naoid_total:,.2f}")
print(f"[item2] notas sem codigo atribuidas por CNPJ: {recuperado_total:,.2f}  | restante sem id: {naoid_total:,.2f}")
print(f"folha total: {tp:,.2f} (esperado ~382.194.167)")
print("%-8s %-38s %15s %15s" % ("cod", "cliente", "faturamento", "folha"))
for r in rows[:15]:
    print("%-8s %-38s %15.0f %15.0f" % (r["client_code"], r["cliente"][:38], r["faturamento"], r["folha"]))

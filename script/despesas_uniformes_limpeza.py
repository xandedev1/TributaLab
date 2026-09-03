#!/usr/bin/env python3
"""Secao Real Audit Tech: despesas de UNIFORMES + MATERIAL DE LIMPEZA (APPA 2025).

Le a planilha 'uniformes e material de limpeza.xlsx' (abas limpeza/uniformes) e
gera um Excel com a base normalizada (com AutoFilter) + abas de analise por:
  - Empresa (codigo de cliente na coluna G; sem codigo => 'SEM EMPRESA RELACIONADA')
  - Data de vencimento (mes)
  - Data de pagamento (mes)
  - Fornecedor

Coluna G ('DESPESA/ CLIENTE') traz o codigo do cliente embutido (inicio/meio/fim).
Casa contra os codigos conhecidos do cruzamento_cliente.json.
"""
import openpyxl, os, re, json, argparse
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INK = "173A3B"; CORAL = "D66E54"; MINT = "9BC8B5"; PAPER = "F5F2E9"; LINE = "D8D5C9"
SEM = "SEM EMPRESA RELACIONADA"
MESNOME = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
           "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]

ap = argparse.ArgumentParser()
ap.add_argument("--src", default=r"C:\Users\xandao\Downloads\uniformes e material de limpeza.xlsx")
ap.add_argument("--mapa", default=r"storage\private\fiscal_auditor\appa\cruzamento_cliente.json")
ap.add_argument("--out", default=r"storage\private\fiscal_auditor\appa\APPA_DESPESAS_UNIFORMES_LIMPEZA_2025.xlsx")
args = ap.parse_args()

# mapa codigo -> nome do cliente (empresa)
cli = json.load(open(args.mapa, encoding="utf-8"))
CODE2NAME = {str(r["client_code"]): r["cliente"] for r in cli["rows"] if r.get("client_code")}
KNOWN = set(CODE2NAME)

REF_RE = re.compile(r"ref\.?\s*\d+", re.I)   # referencias de produto (REF.252) -> nao sao cliente
NUM_RE = re.compile(r"\d{2,4}")

def num(v):
    try:
        return float(str(v).replace(",", ".")) if v not in (None, "") else 0.0
    except Exception:
        return 0.0

def resolve_empresa(g):
    """Extrai o codigo de cliente da descricao G. Retorna (codigo, nome) ou (None, SEM)."""
    txt = str(g or "")
    limpo = REF_RE.sub(" ", txt)             # remove REF.NNN (referencia, nao cliente)
    for tok in NUM_RE.findall(limpo):
        if tok in KNOWN:
            return tok, CODE2NAME[tok]
    return None, SEM

def ym(d):
    return f"{d.year:04d}-{d.month:02d}" if hasattr(d, "year") else None

# ---- le a base ----
wb = openpyxl.load_workbook(args.src, read_only=True, data_only=True)
regs = []
for cat, sn in (("Limpeza", "limpeza"), ("Uniformes", "uniformes")):
    ws = wb[sn]
    first = True
    for r in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if not r or all(x is None for x in r):
            continue
        A, B, C, D, E, F, G, H, I = (list(r) + [None] * 9)[:9]
        cod, emp = resolve_empresa(G)
        regs.append({
            "categoria": cat,
            "venc": A, "pgto": I,
            "fornecedor": re.sub(r"\s*-\s*PRIORIDADE\s*$", "", str(E or "").strip(), flags=re.I),
            "nf": D, "parcela": F, "status": C,
            "cod": cod, "empresa": emp, "descricao": str(G or "").strip(),
            "valor": num(H),
        })
wb.close()

tot = sum(x["valor"] for x in regs)
tot_cat = defaultdict(float)
for x in regs:
    tot_cat[x["categoria"]] += x["valor"]
com = sum(x["valor"] for x in regs if x["cod"])
sem = tot - com
print(f"registros: {len(regs)} | total R$ {tot:,.2f} (limpeza {tot_cat['Limpeza']:,.2f} + uniformes {tot_cat['Uniformes']:,.2f})")
print(f"com empresa: R$ {com:,.2f} ({com/tot*100:.1f}%) | SEM empresa: R$ {sem:,.2f} ({sem/tot*100:.1f}%)")

# ---- estilos ----
f_title = Font(name="Calibri", size=16, bold=True, color=INK)
f_sub = Font(name="Calibri", size=10, color="36595A")
f_hdr = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
f_bold = Font(name="Calibri", size=10, bold=True, color=INK)
f_body = Font(name="Calibri", size=10, color="1A2E2E")
fill_ink = PatternFill("solid", fgColor=INK)
fill_paper = PatternFill("solid", fgColor=PAPER)
fill_coral = PatternFill("solid", fgColor=CORAL)
fill_sem = PatternFill("solid", fgColor="F3D9D2")
thin = Side(style="thin", color=LINE)
border = Border(bottom=thin)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
right = Alignment(horizontal="right", vertical="center")
MONEY = 'R$ #,##0.00'

out_wb = openpyxl.Workbook()

def header_block(ws, title, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, "REAL AUDIT TECH  ·  " + title); c.font = f_title
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(2, 1, sub); c2.font = f_sub
    ws.row_dimensions[1].height = 22

def write_table(ws, start, headers, rows, widths, money_cols, tag_sem_col=None):
    hr = start
    for j, h in enumerate(headers, 1):
        c = ws.cell(hr, j, h); c.font = f_hdr; c.fill = fill_ink
        c.alignment = center if j > 1 else left
    for i, row in enumerate(rows, 1):
        rr = hr + i
        for j, val in enumerate(row, 1):
            c = ws.cell(rr, j, val)
            c.font = f_body
            c.border = border
            if j in money_cols:
                c.number_format = MONEY; c.alignment = right
            elif j == 1:
                c.alignment = left
            else:
                c.alignment = center
            if rr % 2 == 0:
                if not c.fill or c.fill.fgColor.rgb in (None, "00000000"):
                    c.fill = fill_paper
        if tag_sem_col and str(row[tag_sem_col - 1]) == SEM:
            for j in range(1, len(headers) + 1):
                ws.cell(rr, j).fill = fill_sem
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    return hr

def total_row(ws, r, label, vals, money_cols, ncols):
    c = ws.cell(r, 1, label); c.font = Font(bold=True, color="FFFFFF"); c.fill = fill_coral; c.alignment = left
    for j in range(2, ncols + 1):
        c = ws.cell(r, j); c.fill = fill_coral
    for col, v in vals.items():
        c = ws.cell(r, col, v); c.font = Font(bold=True, color="FFFFFF"); c.fill = fill_coral
        c.number_format = MONEY; c.alignment = right

# ===== Aba BASE (filtravel) =====
ws = out_wb.active; ws.title = "Base"
header_block(ws, "Despesas · Uniformes & Material de Limpeza", "APPA · Exercício 2025 · valores pagos (coluna PAGO)", 10)
headers = ["Categoria", "Vencimento", "Pagamento", "Fornecedor", "Nº NF", "Parcela",
           "Cód. Empresa", "Empresa", "Descrição (G)", "Valor Pago"]
base_rows = []
for x in sorted(regs, key=lambda z: (z["categoria"], z["venc"] or 0)):
    base_rows.append([x["categoria"], x["venc"], x["pgto"], x["fornecedor"], x["nf"], x["parcela"],
                      x["cod"] or "—", x["empresa"], x["descricao"], x["valor"]])
hr = write_table(ws, 4, headers, base_rows, [11, 12, 12, 40, 10, 8, 11, 42, 46, 14], {10}, tag_sem_col=8)
for col in (2, 3):
    for r in range(hr + 1, hr + 1 + len(base_rows)):
        ws.cell(r, col).number_format = "dd/mm/yyyy"
ws.auto_filter.ref = f"A{hr}:J{hr + len(base_rows)}"
ws.freeze_panes = f"A{hr + 1}"

# ===== Aba POR EMPRESA =====
ws = out_wb.create_sheet("Por Empresa")
header_block(ws, "Por Empresa (cliente)", "Despesa alocada por código de cliente (coluna G) · SEM EMPRESA = despesa interna/matriz", 6)
agg = defaultdict(lambda: {"Limpeza": 0.0, "Uniformes": 0.0, "n": 0, "cod": ""})
for x in regs:
    k = x["empresa"]
    agg[k][x["categoria"]] += x["valor"]; agg[k]["n"] += 1
    if x["cod"]:
        agg[k]["cod"] = x["cod"]
def sort_emp(kv):
    k, v = kv
    return (k == SEM, -(v["Limpeza"] + v["Uniformes"]))
rows = []
for k, v in sorted(agg.items(), key=sort_emp):
    t = v["Limpeza"] + v["Uniformes"]
    rows.append([k, v["cod"] or "—", v["Limpeza"], v["Uniformes"], t, v["n"]])
hr = write_table(ws, 4, ["Empresa", "Cód.", "Limpeza", "Uniformes", "Total", "Qtd notas"],
                 rows, [46, 8, 16, 16, 16, 10], {3, 4, 5}, tag_sem_col=1)
tr = hr + len(rows) + 1
total_row(ws, tr, "TOTAL", {3: tot_cat["Limpeza"], 4: tot_cat["Uniformes"], 5: tot}, {3, 4, 5}, 6)

# ===== Aba POR VENCIMENTO / POR PAGAMENTO =====
def aba_mes(nome, campo, titulo):
    ws = out_wb.create_sheet(nome)
    header_block(ws, titulo, "APPA · Exercício 2025 · valores pagos", 4)
    agg = defaultdict(lambda: {"Limpeza": 0.0, "Uniformes": 0.0})
    for x in regs:
        k = ym(x[campo])
        if k is None:
            k = "(sem data)"
        agg[k][x["categoria"]] += x["valor"]
    rows = []
    for k in sorted(agg, key=lambda z: (z == "(sem data)", z)):
        v = agg[k]
        label = k if k == "(sem data)" else f"{MESNOME[int(k[5:7])]}/{k[:4]}"
        rows.append([label, v["Limpeza"], v["Uniformes"], v["Limpeza"] + v["Uniformes"]])
    hr = write_table(ws, 4, ["Mês", "Limpeza", "Uniformes", "Total"], rows, [18, 18, 18, 18], {2, 3, 4})
    tr = hr + len(rows) + 1
    total_row(ws, tr, "TOTAL", {2: tot_cat["Limpeza"], 3: tot_cat["Uniformes"], 4: tot}, {2, 3, 4}, 4)

aba_mes("Por Vencimento", "venc", "Por Data de Vencimento")
aba_mes("Por Pagamento", "pgto", "Por Data de Pagamento")

# ===== Aba POR FORNECEDOR =====
ws = out_wb.create_sheet("Por Fornecedor")
header_block(ws, "Por Fornecedor", "APPA · Exercício 2025 · valores pagos", 5)
agg = defaultdict(lambda: {"Limpeza": 0.0, "Uniformes": 0.0, "n": 0})
for x in regs:
    k = x["fornecedor"] or "(sem fornecedor)"
    agg[k][x["categoria"]] += x["valor"]; agg[k]["n"] += 1
rows = []
for k, v in sorted(agg.items(), key=lambda kv: -(kv[1]["Limpeza"] + kv[1]["Uniformes"])):
    rows.append([k, v["Limpeza"], v["Uniformes"], v["Limpeza"] + v["Uniformes"], v["n"]])
hr = write_table(ws, 4, ["Fornecedor", "Limpeza", "Uniformes", "Total", "Qtd notas"],
                 rows, [48, 16, 16, 16, 10], {2, 3, 4})
tr = hr + len(rows) + 1
total_row(ws, tr, "TOTAL", {2: tot_cat["Limpeza"], 3: tot_cat["Uniformes"], 4: tot}, {2, 3, 4}, 5)

# ===== Aba RESUMO =====
ws = out_wb.create_sheet("Resumo", 0)
header_block(ws, "Resumo", "APPA · Despesas de Uniformes & Material de Limpeza · 2025", 2)
kpis = [
    ("Material de Limpeza", tot_cat["Limpeza"]),
    ("Uniformes", tot_cat["Uniformes"]),
    ("TOTAL GERAL", tot),
    ("Alocado a empresa (cliente)", com),
    ("Sem empresa relacionada (interno/matriz)", sem),
]
r = 4
for label, val in kpis:
    c = ws.cell(r, 1, label); c.font = f_bold if "TOTAL" not in label else Font(bold=True, size=12, color=INK)
    c.alignment = left
    c2 = ws.cell(r, 2, val); c2.number_format = MONEY; c2.alignment = right
    c2.font = Font(bold=True, color=INK)
    if "TOTAL GERAL" in label:
        for j in (1, 2):
            ws.cell(r, j).fill = fill_coral; ws.cell(r, j).font = Font(bold=True, color="FFFFFF", size=12)
    r += 1
ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 20

os.makedirs(os.path.dirname(args.out), exist_ok=True)
out_wb.save(args.out)
print("OK ->", args.out)

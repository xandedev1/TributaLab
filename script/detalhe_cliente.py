#!/usr/bin/env python3
"""Detalhe analitico dos clientes deficitarios (>100% prejuizo): faturamento nota a nota
e folha rubrica a rubrica. Escreve um JSON combinado para o gerador de PDFs.
"""
import openpyxl, glob, os, re, json, unicodedata
from collections import defaultdict, Counter

BASE = r"storage\private\fiscal_auditor\appa"
CNPJ_RE = re.compile(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}")
MES = ["01", "02", "03", "04", "05", "06", "07", "08", "09", "10", "11", "12"]

def norm(s):
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()
    return re.sub(r"[^a-z0-9]+", " ", s).strip()

def ident(v):
    if v is None:
        return None
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s or None

def num(v):
    try:
        return float(str(v).replace(",", ".")) if v not in (None, "") else 0.0
    except Exception:
        return 0.0

# alvos = clientes com margem < -100%
cli = json.load(open(os.path.join(BASE, "cruzamento_cliente.json"), encoding="utf-8"))["rows"]
m = json.load(open(os.path.join(BASE, "cruzamento_resultado.json"), encoding="utf-8"))["rows"]
enc = (sum(r["inss_empregador"] for r in m) + sum(r["fgts"] for r in m)) / sum(r["folha_vencimentos"] for r in m)
targets = {}
for r in cli:
    fat = r["faturamento"]
    if fat > 0 and (fat - r["folha"] * (1 + enc)) / fat * 100 < -100:
        targets[r["client_code"]] = {"cliente": r["cliente"], "cnpj": r["cnpj"]}
print("alvos:", {k: v["cliente"] for k, v in targets.items()})

det = {cc: {"cliente": targets[cc]["cliente"], "cnpj": targets[cc]["cnpj"],
            "fat_mes": {mm: {"n": 0, "bruto": 0.0, "ret": 0.0} for mm in MES},
            "notas": [], "folha_rubricas": defaultdict(lambda: {"desc": "", "tipo": "", "mes": [0.0] * 12}),
            "folha_mes_venc": [0.0] * 12, "folha_mes_desc": [0.0] * 12} for cc in targets}

# --- Faturamento nota a nota ---
for f in glob.glob(os.path.join(BASE, "source", "**", "*.xlsx"), recursive=True):
    if "RETEN" not in os.path.basename(f).upper():
        continue
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True); ws = wb.active
    hdr = None
    for row in ws.iter_rows(values_only=True):
        if hdr is None:
            if row and any(isinstance(c, str) and norm(c) == "cnpj cliente" for c in row):
                hdr = {}
                for i, c in enumerate(row):
                    n = norm(c)
                    if n == "cnpj cliente": hdr["cnpj"] = i
                    elif n in ("cod cliente", "cd cliente", "cleinte"): hdr["cc"] = i
                    elif n == "cliente": hdr["cli"] = i
                    elif "competencia" in n: hdr["comp"] = i
                    elif "dt emissao" in n or "data emissao" in n: hdr["emis"] = i
                    elif re.fullmatch(r"n (nf e|nfe)", n): hdr["nf"] = i
                    elif re.fullmatch(r"valor (da )?fatura", n): hdr["bill"] = i
                    elif n in ("valor inss", "valor irrf", "valor pis", "valor cofins", "valor csll", "valor iss"):
                        hdr.setdefault("ret", []).append(i)
            continue
        def g(k):
            i = hdr.get(k)
            return row[i] if i is not None and i < len(row) else None
        cc = ident(g("cc"))
        if cc not in targets:
            continue
        comp = g("comp")
        if not hasattr(comp, "year") or comp.year != 2025:
            continue
        mm = f"{comp.month:02d}"
        bruto = num(g("bill"))
        ret = sum(num(row[i]) for i in hdr.get("ret", []) if i < len(row))
        d = det[cc]
        d["fat_mes"][mm]["n"] += 1
        d["fat_mes"][mm]["bruto"] += bruto
        d["fat_mes"][mm]["ret"] += ret
        emis = g("emis")
        d["notas"].append({"comp": mm, "nf": ident(g("nf")) or "", 
                           "emissao": (emis.strftime("%d/%m/%Y") if hasattr(emis, "year") else ""),
                           "bruto": round(bruto, 2), "ret": round(ret, 2)})
    wb.close()

# --- Folha rubrica a rubrica ---
for f in glob.glob(os.path.join(BASE, "payroll", "*.xlsx")):
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True); ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 17:
            continue
        et = row[4]
        if et not in ("Vencimento", "Desconto"):
            continue
        cc = ident(row[0])
        if cc not in targets:
            continue
        code = ident(row[2]) or ""; desc = str(row[3] or "").strip()
        d = det[cc]
        key = f"{code}|{et}"
        rub = d["folha_rubricas"][key]
        rub["desc"] = desc; rub["tipo"] = et
        for j in range(12):
            v = num(row[5 + j])
            rub["mes"][j] += v
            if et == "Vencimento":
                d["folha_mes_venc"][j] += v
            else:
                d["folha_mes_desc"][j] += v
    wb.close()

# serializa
out = {}
for cc, d in det.items():
    rubricas = []
    for key, rub in d["folha_rubricas"].items():
        code = key.split("|")[0]
        rubricas.append({"codigo": code, "desc": rub["desc"], "tipo": rub["tipo"],
                         "total": round(sum(rub["mes"]), 2), "mes": [round(x, 2) for x in rub["mes"]]})
    rubricas.sort(key=lambda r: (r["tipo"] != "Vencimento", -r["total"]))
    out[cc] = {"cliente": d["cliente"], "cnpj": d["cnpj"],
               "fat_mes": d["fat_mes"], "notas": sorted(d["notas"], key=lambda n: (n["comp"], -n["bruto"])),
               "folha_rubricas": rubricas,
               "folha_mes_venc": [round(x, 2) for x in d["folha_mes_venc"]],
               "folha_mes_desc": [round(x, 2) for x in d["folha_mes_desc"]],
               "encargo": round(enc, 4)}
json.dump(out, open(os.path.join(BASE, "detalhe_clientes_deficit.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)

for cc, d in out.items():
    fat = sum(d["fat_mes"][mm]["bruto"] for mm in MES)
    venc = sum(d["folha_mes_venc"])
    print(f"{d['cliente'][:38]:38} cod={cc:6} notas={len(d['notas']):4} rubricas={len(d['folha_rubricas']):3} fat={fat:14.2f} folha={venc:14.2f}")

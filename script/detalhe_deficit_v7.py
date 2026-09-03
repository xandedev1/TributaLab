#!/usr/bin/env python3
"""Detalhe MEGA por cliente deficitario (>100% prejuizo no modelo v7): busca TUDO no sistema
(faturamento nota a nota, folha rubrica a rubrica com split de beneficios, despesas de
uniformes/limpeza) e a memoria de calculo do %. Escreve JSON para o gerador de PDF.

Modelo v7: Resultado = Faturamento - Folha(sem benef) - Encargos - Beneficios - Mat./Uniformes
"""
import openpyxl, glob, os, re, json, unicodedata
from collections import defaultdict

BASE = r"storage\private\fiscal_auditor\appa"
UNIF = r"C:\Users\xandao\Downloads\uniformes e material de limpeza.xlsx"
DSPREST = r"C:\Users\xandao\Downloads\DSPREST_E_*.xml"
COMPILADO = r"C:\Users\xandao\Downloads\Compilado_NFs_2026.xlsx"
CNPJ_RE = re.compile(r"\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}")
MES = [f"{k:02d}" for k in range(1, 13)]
BEN_KW = ("VALE", "REFEI", "ALIMENT", "CESTA", "MEDIC", "ODONT", "LANCHE", "TRANSP")
REF = re.compile(r"ref\.?\s*\d+", re.I)
NUM = re.compile(r"\d{2,4}")

def norm(s):
    return re.sub(r"[^a-z0-9]+", " ", unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()).strip()

def ident(v):
    if v is None:
        return None
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s or None

def num(v):
    try:
        return float(str(v).replace(",", ".")) if v not in (None, "") else 0.0
    except Exception:
        return 0.0

def is_benef(desc):
    dn = norm(desc)
    if any(x in dn for x in ("pensao", "contrib", "sindic", "assistencial")):
        return False
    return any(k.lower() in dn for k in BEN_KW)

cli = json.load(open(os.path.join(BASE, "cruzamento_cliente.json"), encoding="utf-8"))
rows = cli["rows"]
men = json.load(open(os.path.join(BASE, "cruzamento_resultado.json"), encoding="utf-8"))["rows"]
inss = sum(r["inss_empregador"] for r in men); fgts = sum(r["fgts"] for r in men)
folha_m = sum(r["folha_vencimentos"] for r in men); ben_add_tot = cli.get("total_beneficio_add", 0.0)
enc = (inss + fgts) / (folha_m - ben_add_tot)
base_fat = sum(r["faturamento"] for r in rows)
code2name = {str(r["client_code"]): r["cliente"] for r in rows if r.get("client_code")}
known = set(code2name)

# despesas de uniformes/limpeza por codigo + total sem-empresa (para rateio)
unif_direto = defaultdict(float); unif_sem = 0.0
unif_notas = defaultdict(list)
wb = openpyxl.load_workbook(UNIF, read_only=True, data_only=True)
for cat, sn in (("Limpeza", "limpeza"), ("Uniformes", "uniformes")):
    ws = wb[sn]; first = True
    for r in ws.iter_rows(values_only=True):
        if first:
            first = False; continue
        if not r or all(x is None for x in r):
            continue
        row = list(r) + [None] * 9
        v = num(row[7])
        cod = next((t for t in NUM.findall(REF.sub(" ", str(row[6] or ""))) if t in known), None)
        if cod:
            unif_direto[cod] += v
            unif_notas[cod].append({"cat": cat, "fornecedor": str(row[4] or "").strip(),
                                    "desc": str(row[6] or "").strip(), "valor": round(v, 2)})
        else:
            unif_sem += v
wb.close()

# alvos = margem v7 < -100%
targets = {}
for r in rows:
    fat = r["faturamento"]
    if fat <= 0:
        continue
    fol = r["folha_sem_beneficio"]; ben = r["beneficio_liquido"]
    uni = unif_direto.get(str(r["client_code"]), 0.0) + unif_sem * (fat / base_fat)
    custo = fol + fol * enc + ben + uni
    if (fat - custo) / fat * 100 < -100:
        targets[str(r["client_code"])] = r
print("alvos:", {k: v["cliente"][:34] for k, v in targets.items()})

cnpj2cc = {}
for cc, r in targets.items():
    if r.get("cnpj"):
        cnpj2cc[re.sub(r"\D", "", r["cnpj"])] = cc

det = {cc: {"cliente": r["cliente"], "cnpj": r.get("cnpj", ""), "codigo": cc,
            "faturamento": r["faturamento"], "folha_sem_beneficio": r["folha_sem_beneficio"],
            "beneficio_liquido": r["beneficio_liquido"], "beneficio_add": r.get("beneficio_add", 0.0),
            "beneficio_sub": r.get("beneficio_sub", 0.0), "folha_bruta": r["folha"],
            "fat_mes": {mm: {"n": 0, "bruto": 0.0, "ret": 0.0} for mm in MES}, "notas": [],
            "folha_rubricas": defaultdict(lambda: {"desc": "", "tipo": "", "benef": False, "mes": [0.0] * 12}),
            "folha_mes_venc": [0.0] * 12, "folha_mes_desc": [0.0] * 12}
       for cc, r in targets.items()}
seen = set(); uniq = [0]

def add_nota(cc, mm, nf, emis, bruto, ret, fonte):
    d = det[cc]
    d["fat_mes"][mm]["n"] += 1; d["fat_mes"][mm]["bruto"] += bruto; d["fat_mes"][mm]["ret"] += ret
    d["notas"].append({"comp": mm, "nf": nf or "", "emissao": emis, "bruto": round(bruto, 2),
                       "ret": round(ret, 2), "fonte": fonte})

# --- Faturamento nota a nota (RETEN 2025/2026 + Compilado) ---
fontes = [f for f in glob.glob(os.path.join(BASE, "source", "**", "*.xlsx"), recursive=True)
          if "RETEN" in os.path.basename(f).upper()]
if os.path.exists(COMPILADO):
    fontes.append(COMPILADO)
for f in fontes:
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True); ws = wb.active
    hdr = None
    for row in ws.iter_rows(values_only=True):
        if hdr is None:
            if row and any(isinstance(c, str) and norm(c) == "cnpj cliente" for c in row):
                hdr = {}
                for i, c in enumerate(row):
                    n = norm(c)
                    if n == "cnpj cliente": hdr["cnpj"] = i
                    elif n in ("cod cliente", "cd cliente", "cleinte"): hdr["cc"] = i
                    elif n == "cliente": hdr["cli"] = i
                    elif "competencia" in n: hdr["comp"] = i
                    elif "emissao" in n: hdr["emis"] = i
                    elif re.fullmatch(r"no? nf ?e", n): hdr["nf"] = i
                    elif re.fullmatch(r"valor (da )?fatura", n): hdr["bill"] = i
                    elif n in ("valor inss", "valor irrf", "valor pis", "valor cofins", "valor csll", "valor iss"):
                        hdr.setdefault("ret", []).append(i)
            continue
        def g(k):
            i = hdr.get(k)
            return row[i] if i is not None and i < len(row) else None
        cnpj = str(g("cnpj") or "").strip()
        cc = ident(g("cc"))
        # resolve alvo: por codigo, ou por CNPJ do alvo (item 2)
        tcc = cc if cc in targets else cnpj2cc.get(re.sub(r"\D", "", cnpj))
        if tcc not in targets:
            continue
        comp = g("comp")
        if not hasattr(comp, "year") or comp.year != 2025:
            continue
        nf = ident(g("nf"))
        key = (cc or cnpj, nf) if nf else ("_", uniq[0])
        if not nf:
            uniq[0] += 1
        if key in seen:
            continue
        seen.add(key)
        bruto = num(g("bill")); ret = sum(num(row[i]) for i in hdr.get("ret", []) if i < len(row))
        emis = g("emis")
        add_nota(tcc, f"{comp.month:02d}", nf, emis.strftime("%d/%m/%Y") if hasattr(emis, "year") else "",
                 bruto, ret, os.path.basename(f)[:22])
    wb.close()

# --- DSPREST (NFS-e) para os alvos (ex.: CAIXA NF142) ---
MESES_N = {"JANEIRO": 1, "FEVEREIRO": 2, "MARCO": 3, "ABRIL": 4, "MAIO": 5, "JUNHO": 6,
           "JULHO": 7, "AGOSTO": 8, "SETEMBRO": 9, "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12}
for f in glob.glob(DSPREST):
    d = open(f, encoding="utf-8", errors="replace").read()
    tom = re.search(r"<TomadorServico>.*?<Cnpj>(\d+)</Cnpj>", d, re.S)
    if not tom:
        continue
    tcc = cnpj2cc.get(tom.group(1))
    if tcc not in targets:
        continue
    mm = re.search(r"MES DE (\w+)", d)
    mi = MESES_N.get(norm(mm.group(1)).upper()) if mm else None
    if mi is None:
        continue
    nfm = re.search(r"<Numero>(\d+)</Numero>", d)
    nf = ident(nfm.group(1)) if nfm else None
    key = (tcc, nf) if nf else ("_dsp", uniq[0])
    if not nf:
        uniq[0] += 1
    if key in seen:
        continue
    seen.add(key)
    vb = re.search(r"<ValorServicos>([\d.]+)</ValorServicos>", d)
    bruto = float(vb.group(1)) if vb else 0.0
    add_nota(tcc, f"{mi:02d}", nf, "", bruto, 0.0, "DSPREST NFS-e")

# --- Folha rubrica a rubrica (split beneficio) ---
for f in glob.glob(os.path.join(BASE, "payroll", "*.xlsx")):
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True); ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 17:
            continue
        et = row[4]
        if et not in ("Vencimento", "Desconto"):
            continue
        cc = ident(row[0])
        if cc not in targets:
            continue
        code = ident(row[2]) or ""; desc = str(row[3] or "").strip()
        d = det[cc]; benef = is_benef(desc)
        rub = d["folha_rubricas"][f"{code}|{et}"]
        rub["desc"] = desc; rub["tipo"] = et; rub["benef"] = benef
        for j in range(12):
            v = num(row[5 + j])
            rub["mes"][j] += v
            if et == "Vencimento":
                d["folha_mes_venc"][j] += v
            else:
                d["folha_mes_desc"][j] += v
    wb.close()

out = {}
for cc, d in det.items():
    rubricas = []
    for key, rub in d["folha_rubricas"].items():
        rubricas.append({"codigo": key.split("|")[0], "desc": rub["desc"], "tipo": rub["tipo"],
                         "benef": rub["benef"], "total": round(sum(rub["mes"]), 2)})
    rubricas.sort(key=lambda r: (r["tipo"] != "Vencimento", -r["total"]))
    fat = d["faturamento"]; fol = d["folha_sem_beneficio"]; ben = d["beneficio_liquido"]
    uni_dir = round(unif_direto.get(cc, 0.0), 2); uni_rat = round(unif_sem * (fat / base_fat), 2)
    uni = uni_dir + uni_rat; encv = round(fol * enc, 2)
    resultado = round(fat - fol - encv - ben - uni, 2)
    out[cc] = {**{k: d[k] for k in ("cliente", "cnpj", "codigo", "faturamento", "folha_sem_beneficio",
                                    "beneficio_liquido", "beneficio_add", "beneficio_sub", "folha_bruta",
                                    "fat_mes", "folha_mes_venc", "folha_mes_desc")},
               "notas": sorted(d["notas"], key=lambda n: (n["comp"], -n["bruto"])),
               "folha_rubricas": rubricas,
               "uniformes_notas": sorted(unif_notas.get(cc, []), key=lambda n: -n["valor"]),
               "uniformes_direto": uni_dir, "uniformes_rateio": uni_rat, "uniformes_total": round(uni, 2),
               "encargo": round(enc, 4), "encargo_valor": encv,
               "resultado": resultado, "margem": round(resultado / fat * 100, 1) if fat else 0.0}
json.dump(out, open(os.path.join(BASE, "detalhe_deficit_v7.json"), "w", encoding="utf-8"), ensure_ascii=False, indent=1)

for cc, d in out.items():
    print(f"{d['cliente'][:34]:34} cod={cc:5} notas={len(d['notas']):3} fat={d['faturamento']:13,.2f} "
          f"folha={d['folha_sem_beneficio']:13,.2f} unif={d['uniformes_total']:11,.2f} margem={d['margem']:7.0f}%")

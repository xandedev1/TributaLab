import datetime
import json
import sys
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from pyxlsb import open_workbook


def serialize(value):
    if value is None:
        return ""
    if isinstance(value, (datetime.date, datetime.datetime, datetime.time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value)


def xlsx_view(path, requested_sheet, start_row, page_size):
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if requested_sheet and requested_sheet not in workbook.sheetnames:
            raise ValueError("Worksheet not found")
        sheet_name = requested_sheet or workbook.sheetnames[0]
        sheet = workbook[sheet_name]
        total_rows = sheet.max_row
        total_columns = sheet.max_column
        end_row = min(start_row + page_size - 1, total_rows)
        rows = [
            {
                "number": row_number,
                "cells": [serialize(cell.value) for cell in row],
            }
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=start_row, max_row=end_row, max_col=total_columns),
                start=start_row,
            )
        ]
        return workbook.sheetnames, sheet_name, total_rows, total_columns, rows
    finally:
        workbook.close()


def xlsb_view(path, requested_sheet, start_row, page_size):
    with open_workbook(path) as workbook:
        if requested_sheet and requested_sheet not in workbook.sheets:
            raise ValueError("Worksheet not found")
        sheet_name = requested_sheet or workbook.sheets[0]
        rows = []
        total_rows = 0
        total_columns = 0
        end_row = start_row + page_size - 1
        with workbook.get_sheet(sheet_name) as sheet:
            for row_number, row in enumerate(sheet.rows(), start=1):
                values = [serialize(cell.v) for cell in row]
                total_rows = row_number
                total_columns = max(total_columns, len(values))
                if start_row <= row_number <= end_row:
                    rows.append({"number": row_number, "cells": values})

        for row in rows:
            row["cells"].extend([""] * (total_columns - len(row["cells"])))
        return workbook.sheets, sheet_name, total_rows, total_columns, rows


def extract(path, requested_sheet, start_row, page_size):
    extension = Path(path).suffix.lower()
    if extension == ".xlsx":
        sheets, sheet, total_rows, total_columns, rows = xlsx_view(
            path, requested_sheet, start_row, page_size
        )
    elif extension == ".xlsb":
        sheets, sheet, total_rows, total_columns, rows = xlsb_view(
            path, requested_sheet, start_row, page_size
        )
    else:
        raise ValueError("Unsupported spreadsheet format")

    return {
        "filename": Path(path).name,
        "sheet": sheet,
        "sheets": sheets,
        "total_rows": total_rows,
        "total_columns": total_columns,
        "rows": rows,
    }


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    result = extract(
        sys.argv[1],
        sys.argv[2] or None,
        max(int(sys.argv[3]), 1),
        max(int(sys.argv[4]), 1),
    )
    json.dump(result, sys.stdout, ensure_ascii=False, separators=(",", ":"))
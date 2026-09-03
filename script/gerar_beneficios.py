#!/usr/bin/env python3
"""Tabela Geral de Beneficios (APPA 2025): aditivos (empresa) - subtrativos (empregado) = valor final.
Gera Excel (resumo + detalhe + reconciliacao) e PDF. Valores da folha (soma anual por evento)."""
import glob, os, unicodedata
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (BaseDocTemplate, PageTemplate, Frame, Paragraph, Spacer,
                                Table, TableStyle, HRFlowable)

PAY = r"storage\private\fiscal_auditor\appa\payroll"
XLS = r"C:\Users\xandao\Downloads\APPA_Tabela_Geral_Beneficios_2025.xlsx"
PDF = r"C:\Users\xandao\Downloads\APPA_TABELA_GERAL_BENEFICIOS_2025.pdf"

# Mapeamento do print (aditivo corrigido p/ o codigo que carrega valor na folha; add_print = codigo do print)
BEN = [
    ("Plano Médico", "9279", "9279", ["537", "607", "774", "779"]),
    ("Plano Odontológico", "9281", "9281", ["631", "765", "775", "894"]),
    ("Vale-Transporte", "8276", "9276", ["576", "779"]),
    ("Refeição", "8277", "9277", ["531", "623", "624", "702", "727", "728", "769", "770", "771", "773", "1018"]),
    ("Alimentação", "8281", "9284", ["568", "577", "578", "611", "623", "664", "710", "776", "778"]),
    ("Cesta Básica", "8278", "9278", ["772", "1148"]),
]
KW_BEN = [("ODONT", "Plano Odontológico"), ("MEDIC", "Plano Médico"), ("MÉDIC", "Plano Médico"),
          ("TRANSP", "Vale-Transporte"), ("REFEI", "Refeição"), ("LANCHE", "Refeição"),
          ("ALIMENT", "Alimentação"), ("CESTA", "Cesta Básica")]

def nrm(s):
    return unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().upper()

def num(v):
    try:
        return float(str(v).replace(",", ".")) if v not in (None, "") else 0.0
    except Exception:
        return 0.0

# agrega todos os eventos da folha
ev = {}
for f in glob.glob(os.path.join(PAY, "*.xlsx")):
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True); ws = wb.active
    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 17:
            continue
        code = str(row[2]).strip().rstrip(".0") if row[2] is not None else ""
        if not code:
            continue
        et = str(row[4] or "").strip()
        d = ev.setdefault(code, {"desc": "", "tipo": "", "val": 0.0})
        if row[3]:
            d["desc"] = str(row[3]).strip()
        if et:
            d["tipo"] = et
        d["val"] += sum(num(row[c]) for c in range(5, 17))
    wb.close()

def val(c):
    return ev.get(c, {}).get("val", 0.0)
def desc(c):
    return ev.get(c, {}).get("desc", "")
def tipo(c):
    return ev.get(c, {}).get("tipo", "")

# monta beneficios
usados = set()
beneficios = []
for nome, add, add_print, subs in BEN:
    usados.add(add); usados.update(subs); usados.add(add_print)
    av = val(add); sv = sum(val(s) for s in subs)
    beneficios.append({
        "nome": nome, "add": add, "add_print": add_print, "add_desc": desc(add), "add_val": av,
        "subs": subs, "sub_val": sv, "final": av - sv,
    })

# ==================== EXCEL ====================
wb = openpyxl.Workbook()
INK = "173A3B"; CORAL = "D66E54"
hf = Font(bold=True, color="FFFFFF"); hfill = PatternFill("solid", fgColor=INK)
money = "#,##0.00"; thin = Side(style="thin", color="D8D5C9")
def head(ws, cols):
    for j, (n, w) in enumerate(cols, 1):
        c = ws.cell(1, j, n); c.font = hf; c.fill = hfill
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
    ws.freeze_panes = "A2"

# aba resumo
ws = wb.active; ws.title = "Tabela Geral"
head(ws, [("Benefício", 22), ("Cód. Aditivo", 12), ("Aditivo (empresa) R$", 20),
          ("Subtrativo (empregado) R$", 22), ("Valor Final R$", 18)])
r = 2
for b in beneficios:
    ws.cell(r, 1, b["nome"]); ws.cell(r, 2, b["add"])
    ws.cell(r, 3, round(b["add_val"], 2)).number_format = money
    ws.cell(r, 4, round(b["sub_val"], 2)).number_format = money
    fc = ws.cell(r, 5, round(b["final"], 2)); fc.number_format = money; fc.font = Font(bold=True)
    r += 1
ws.cell(r, 1, "TOTAL").font = Font(bold=True)
for col, key in ((3, "add_val"), (4, "sub_val"), (5, "final")):
    c = ws.cell(r, col, round(sum(b[key] for b in beneficios), 2)); c.number_format = money; c.font = Font(bold=True)

# aba detalhe por evento
ws2 = wb.create_sheet("Detalhe por Evento")
head(ws2, [("Benefício", 22), ("Tipo", 12), ("Cód. Evento", 12), ("Descrição", 44), ("Tipo (folha)", 14), ("Valor R$", 16)])
r = 2
for b in beneficios:
    ws2.cell(r, 1, b["nome"]); ws2.cell(r, 2, "Aditivo (+)"); ws2.cell(r, 3, b["add"])
    ws2.cell(r, 4, b["add_desc"]); ws2.cell(r, 5, tipo(b["add"]))
    ws2.cell(r, 6, round(b["add_val"], 2)).number_format = money
    r += 1
    for s in b["subs"]:
        ws2.cell(r, 1, b["nome"]); ws2.cell(r, 2, "Subtrativo (−)"); ws2.cell(r, 3, s)
        ws2.cell(r, 4, desc(s)); ws2.cell(r, 5, tipo(s))
        ws2.cell(r, 6, round(val(s), 2)).number_format = money
        r += 1
    fc = ws2.cell(r, 1, f"= Valor final {b['nome']}"); fc.font = Font(bold=True)
    c = ws2.cell(r, 6, round(b["final"], 2)); c.number_format = money; c.font = Font(bold=True, color=CORAL)
    r += 2

# aba reconciliacao: codigos de beneficio na folha fora do mapeamento
ws3 = wb.create_sheet("Reconciliação (fora do print)")
head(ws3, [("Cód. Evento", 12), ("Descrição", 46), ("Tipo (folha)", 14), ("Benefício sugerido", 20), ("Valor R$", 16)])
r = 2
extra = []
for code, d in ev.items():
    if code in usados or d["val"] <= 0:
        continue
    dn = nrm(d["desc"])
    sug = next((b for k, b in KW_BEN if k in dn), None)
    if sug:
        extra.append((code, d["desc"], d["tipo"], sug, d["val"]))
for code, ds, tp, sug, v in sorted(extra, key=lambda x: -x[4]):
    ws3.cell(r, 1, code); ws3.cell(r, 2, ds); ws3.cell(r, 3, tp); ws3.cell(r, 4, sug)
    ws3.cell(r, 5, round(v, 2)).number_format = money
    r += 1
wb.save(XLS)

# ==================== PDF ====================
def brl(v):
    s = f"{abs(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return ("-" if v < 0 else "") + s
INKc = colors.HexColor("#173a3b"); CORALc = colors.HexColor("#d66e54"); MINT = colors.HexColor("#9bc8b5")
PAPER = colors.HexColor("#f5f2e9"); LINE = colors.HexColor("#d8d5c9"); GREEN = colors.HexColor("#2e7d54")

def hfp(canvas, doc):
    canvas.saveState(); w, h = A4
    canvas.setFillColor(INKc); canvas.rect(0, h - 16 * mm, w, 16 * mm, fill=1, stroke=0)
    canvas.setFillColor(CORALc); canvas.rect(0, h - 16 * mm, 2.6 * mm, 16 * mm, fill=1, stroke=0)
    canvas.setFillColor(CORALc); canvas.rect(18 * mm, h - 12.9 * mm, 6.6 * mm, 6.6 * mm, fill=1, stroke=0)
    canvas.setFillColor(colors.white); canvas.setFont("Helvetica-Bold", 6.6)
    canvas.drawCentredString(18 * mm + 3.3 * mm, h - 11.0 * mm, "RAT")
    canvas.setFont("Helvetica-Bold", 10); canvas.drawString(27.5 * mm, h - 9.8 * mm, "REAL AUDIT TECH")
    canvas.setFont("Helvetica", 5.8); canvas.setFillColor(MINT); canvas.drawString(27.7 * mm, h - 13.2 * mm, "TECNOLOGIA TRIBUTÁRIA")
    canvas.setFont("Helvetica", 8); canvas.setFillColor(MINT); canvas.drawRightString(w - 18 * mm, h - 10.5 * mm, f"Benefícios · APPA 2025 · Pág. {doc.page}")
    canvas.setStrokeColor(LINE); canvas.setLineWidth(0.5); canvas.line(18 * mm, 12 * mm, w - 18 * mm, 12 * mm)
    canvas.setFillColor(colors.HexColor("#738080")); canvas.setFont("Helvetica", 6.5)
    canvas.drawString(18 * mm, 8.5 * mm, "Uso restrito · documento gerencial. Benefícios: parte da empresa (aditivo) menos parte do empregado (subtrativo).")
    canvas.drawRightString(w - 18 * mm, 8.5 * mm, date.today().strftime("%d/%m/%Y"))
    canvas.restoreState()

S = getSampleStyleSheet()
h1 = ParagraphStyle("h1", parent=S["Title"], textColor=INKc, fontSize=22, leading=25, alignment=0)
h1s = ParagraphStyle("h1s", parent=S["Title"], textColor=CORALc, fontSize=13, leading=15, alignment=0)
kick = ParagraphStyle("kick", parent=S["Normal"], textColor=CORALc, fontSize=9, fontName="Helvetica-Bold", spaceAfter=8)
body = ParagraphStyle("body", parent=S["Normal"], textColor=colors.HexColor("#0c292a"), fontSize=8.6, leading=13)
sec = ParagraphStyle("sec", parent=S["Heading2"], textColor=INKc, fontSize=13, leading=16, spaceBefore=10, spaceAfter=5)
cell = ParagraphStyle("cell", parent=body, fontSize=7.2, leading=8.6)

doc = BaseDocTemplate(PDF, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm, topMargin=22 * mm, bottomMargin=16 * mm)
doc.addPageTemplates([PageTemplate(id="m", frames=[Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height)], onPage=hfp)])
el = [Spacer(1, 4)]
el.append(Paragraph("RELATÓRIO GERENCIAL · BENEFÍCIOS", kick))
el.append(Paragraph("Tabela Geral de Benefícios", h1))
el.append(Paragraph("APPA · Exercício 2025", h1s))
el.append(HRFlowable(width="100%", thickness=1.5, color=CORALc, spaceBefore=6, spaceAfter=8))
el.append(Paragraph("Cada benefício é composto pela <b>parte da empresa (aditivo, +)</b> menos a <b>parte do "
                    "empregado (subtrativo, −)</b>, resultando no <b>valor final</b> (custo líquido para a empresa). "
                    "Valores somados da folha de pagamento de 2025.", body))
el.append(Spacer(1, 8))

hdr = ["Benefício", "Aditivo — empresa (+)", "Subtrativo — empregado (−)", "Valor final"]
data = [hdr]
for b in beneficios:
    data.append([b["nome"], brl(b["add_val"]), brl(b["sub_val"]), brl(b["final"])])
data.append(["TOTAL", brl(sum(b["add_val"] for b in beneficios)), brl(sum(b["sub_val"] for b in beneficios)), brl(sum(b["final"] for b in beneficios))])
t = Table(data, colWidths=[42 * mm, 44 * mm, 48 * mm, 40 * mm], repeatRows=1)
t.setStyle(TableStyle([
    ("BACKGROUND", (0, 0), (-1, 0), INKc), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 8.4),
    ("ALIGN", (1, 0), (-1, -1), "RIGHT"), ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, PAPER]),
    ("TEXTCOLOR", (3, 1), (3, -1), GREEN), ("FONTNAME", (3, 1), (3, -1), "Helvetica-Bold"),
    ("LINEABOVE", (0, -1), (-1, -1), 0.7, INKc), ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#eef3ef")),
    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4), ("LEFTPADDING", (0, 0), (-1, -1), 6)]))
el.append(t)
el.append(Spacer(1, 5))
el.append(Paragraph(f"<b>Custo líquido total de benefícios (2025): R$ {brl(sum(b['final'] for b in beneficios))}</b> "
                    "(parte da empresa menos a parte descontada dos empregados).", body))
el.append(Spacer(1, 10))

el.append(Paragraph("Detalhe por evento", sec))
d2 = [["Benefício", "Tipo", "Cód.", "Descrição", "Valor R$"]]
for b in beneficios:
    d2.append([b["nome"], "Aditivo (+)", b["add"], Paragraph(b["add_desc"] or "—", cell), brl(b["add_val"])])
    for s in b["subs"]:
        d2.append(["", "Subtrativo (−)", s, Paragraph(desc(s) or "—", cell), brl(val(s))])
    d2.append(["", Paragraph("<b>= valor final</b>", cell), "", "", brl(b["final"])])
t2 = Table(d2, colWidths=[30 * mm, 24 * mm, 12 * mm, 82 * mm, 26 * mm], repeatRows=1)
stt = [("BACKGROUND", (0, 0), (-1, 0), INKc), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
       ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 6.9),
       ("ALIGN", (4, 0), (4, -1), "RIGHT"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
       ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, PAPER]),
       ("TOPPADDING", (0, 0), (-1, -1), 2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2), ("LEFTPADDING", (0, 0), (-1, -1), 4)]
el.append(Table(d2, colWidths=[30 * mm, 24 * mm, 12 * mm, 82 * mm, 26 * mm], repeatRows=1, style=TableStyle(stt)))
el.append(Spacer(1, 8))

el.append(Paragraph("Notas de reconciliação", sec))
el.append(Paragraph(
    "1) Os aditivos (parte da empresa) de VT, refeição, alimentação e cesta foram lidos na <b>série 8xxx</b> da folha "
    "(8276 VT, 8277 refeição, 8281 alimentação, 8278 cesta), pois os códigos 9276/9277/9278/9284 do print estão "
    "<b>zerados</b> na folha (são informativos eSocial). Médico (9279) e odontológico (9281) conferem com o print.<br/>"
    "2) A lista de subtrativos do print não cobre todos os descontos: há códigos relevantes fora do mapeamento — "
    f"o maior é o <b>672 “Desc. Vale-Transporte” = R$ {brl(val('672'))}</b>. Todos os descontos de benefício não "
    "mapeados estão na aba <b>“Reconciliação”</b> do Excel, com o benefício sugerido, para sua validação.", body))
doc.build(el)

print("OK Excel ->", XLS)
print("OK PDF   ->", PDF)
print("Tabela geral:")
for b in beneficios:
    print(f"  {b['nome']:20} add({b['add']})={b['add_val']:>14,.2f}  sub={b['sub_val']:>13,.2f}  final={b['final']:>14,.2f}")

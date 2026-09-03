#!/usr/bin/env python3
"""Gera um Excel com TODAS as rubricas do S-1010 (tabela de rubricas eSocial) da APPA,
com seus atributos: descricao, natureza, tipo e incidencias CP/IRRF/FGTS/SIND."""
import glob, os, zipfile, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = r"C:\Users\xandao\Downloads\appa tabela gerais\S1010 TODOS ANOS APPA"
OUT = r"C:\Users\xandao\Downloads\APPA_Rubricas_S1010_todos_anos.xlsx"

def g(tag, s):
    m = re.search(rf"<{tag}>(.*?)</{tag}>", s)
    return m.group(1).strip() if m else ""

INCCP = {"00": "Não é base CP", "11": "Base CP mensal", "12": "Base CP 13º",
         "21": "Exposição/aposent. especial", "91": "Suspensa judicial (mensal)",
         "92": "Suspensa judicial (13º)", "93": "Suspensa (exposição)",
         "94": "Suspensa", "95": "Suspensa judicial", "96": "Suspensa", "97": "Suspensa", "98": "Suspensa", "99": "Suspensa"}

registros = []
op_re = re.compile(r"<(inclusao|alteracao|exclusao)>([\s\S]*?)</\1>")
evt_re = re.compile(r'<evtTabRubrica Id="([^"]+)">([\s\S]*?)</evtTabRubrica>')
zips = sorted(glob.glob(os.path.join(BASE, "**", "*.zip"), recursive=True))
for z in zips:
    ano = os.path.basename(os.path.dirname(z))
    try:
        with zipfile.ZipFile(z) as zf:
            for n in zf.namelist():
                if not n.lower().endswith(".xml"):
                    continue
                d = zf.read(n).decode("utf-8", "replace")
                if "evtTabRubrica" not in d:
                    continue
                for eid, ebody in evt_re.findall(d):
                    for op, blk in op_re.findall(ebody):
                        proc = re.search(r"<ideProcessoCP>([\s\S]*?)</ideProcessoCP>", blk)
                        registros.append({
                            "codRubr": g("codRubr", blk),
                            "ideTabRubr": g("ideTabRubr", blk),
                            "iniValid": g("iniValid", blk),
                            "fimValid": g("fimValid", blk),
                            "operacao": op,
                            "dscRubr": g("dscRubr", blk),
                            "natRubr": g("natRubr", blk),
                            "tpRubr": g("tpRubr", blk),
                            "codIncCP": g("codIncCP", blk),
                            "codIncCP_desc": INCCP.get(g("codIncCP", blk), ""),
                            "codIncIRRF": g("codIncIRRF", blk),
                            "codIncFGTS": g("codIncFGTS", blk),
                            "codIncSIND": g("codIncSIND", blk),
                            "processoCP": (g("nrProc", proc.group(1)) if proc else ""),
                            "codSusp": (g("codSusp", proc.group(1)) if proc else ""),
                            "ano_arquivo": ano,
                            "evento_Id": eid,
                        })
    except Exception as e:
        print("erro em", z, e)

# ordena por codRubr (numerico quando possivel), ideTabRubr, iniValid
def keyf(r):
    try:
        c = int(r["codRubr"])
    except Exception:
        c = 10**9
    return (r["ideTabRubr"], c, r["codRubr"], r["iniValid"])
registros.sort(key=keyf)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Rubricas S-1010"
cols = [("codRubr", 10), ("ideTabRubr", 10), ("iniValid", 10), ("fimValid", 10), ("operacao", 11),
        ("dscRubr", 46), ("natRubr", 9), ("tpRubr", 8), ("codIncCP", 9), ("codIncCP_desc", 24),
        ("codIncIRRF", 10), ("codIncFGTS", 11), ("codIncSIND", 11), ("processoCP", 22), ("codSusp", 9),
        ("ano_arquivo", 12), ("evento_Id", 40)]
INK = "173A3B"; head_font = Font(bold=True, color="FFFFFF"); head_fill = PatternFill("solid", fgColor=INK)
thin = Side(style="thin", color="D8D5C9")
for j, (name, w) in enumerate(cols, 1):
    c = ws.cell(1, j, name)
    c.font = head_font; c.fill = head_fill; c.alignment = Alignment(vertical="center")
    ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
for i, r in enumerate(registros, 2):
    for j, (name, _) in enumerate(cols, 1):
        cell = ws.cell(i, j, r[name])
        cell.border = Border(bottom=thin)
    if i % 2 == 0:
        for j in range(1, len(cols) + 1):
            ws.cell(i, j).fill = PatternFill("solid", fgColor="F5F2E9")
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(cols))}{len(registros)+1}"

# aba resumo: rubricas distintas
distintas = {}
for r in registros:
    distintas.setdefault((r["ideTabRubr"], r["codRubr"]), r)
ws2 = wb.create_sheet("Resumo")
ws2["A1"] = "Total de registros (inclusão/alteração/exclusão)"; ws2["B1"] = len(registros)
ws2["A2"] = "Rubricas distintas (ideTabRubr + codRubr)"; ws2["B2"] = len(distintas)
ws2["A3"] = "Tabelas de rubrica (ideTabRubr)"; ws2["B3"] = ", ".join(sorted({r["ideTabRubr"] for r in registros}))
ws2["A4"] = "Zips processados"; ws2["B4"] = len(zips)
for r in range(1, 5):
    ws2.cell(r, 1).font = Font(bold=True)
ws2.column_dimensions["A"].width = 48; ws2.column_dimensions["B"].width = 40

wb.save(OUT)
print(f"OK -> {OUT}")
print(f"registros: {len(registros)} | rubricas distintas: {len(distintas)} | zips: {len(zips)}")

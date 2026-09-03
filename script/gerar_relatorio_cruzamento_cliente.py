#!/usr/bin/env python3
"""Cruzamento por Cliente (tomador) — Faturamento x Folha 2025 (PDF, Real Audit Tech)."""
import argparse, json, os, re
from datetime import date
import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (BaseDocTemplate, PageTemplate, Frame, Paragraph, Spacer,
                                Table, TableStyle, HRFlowable)

INK = colors.HexColor("#173a3b"); INK2 = colors.HexColor("#0c292a")
CORAL = colors.HexColor("#d66e54"); MINT = colors.HexColor("#9bc8b5")
PAPER = colors.HexColor("#f5f2e9"); LINE = colors.HexColor("#d8d5c9")
GREEN = colors.HexColor("#2e7d54"); RED = colors.HexColor("#c0492f")

def brl(v):
    if v is None:
        return "—"
    s = f"{abs(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return ("-" if v < 0 else "") + s

def header_footer(canvas, doc):
    canvas.saveState()
    w, h = A4
    canvas.setFillColor(INK); canvas.rect(0, h - 16 * mm, w, 16 * mm, fill=1, stroke=0)
    canvas.setFillColor(CORAL); canvas.rect(0, h - 16 * mm, 2.6 * mm, 16 * mm, fill=1, stroke=0)
    canvas.setFillColor(CORAL); canvas.rect(18 * mm, h - 12.9 * mm, 6.6 * mm, 6.6 * mm, fill=1, stroke=0)
    canvas.setFillColor(colors.white); canvas.setFont("Helvetica-Bold", 6.6)
    canvas.drawCentredString(18 * mm + 3.3 * mm, h - 11.0 * mm, "RAT")
    canvas.setFont("Helvetica-Bold", 10)
    canvas.drawString(27.5 * mm, h - 9.8 * mm, "REAL AUDIT TECH")
    canvas.setFont("Helvetica", 5.8); canvas.setFillColor(MINT)
    canvas.drawString(27.7 * mm, h - 13.2 * mm, "TECNOLOGIA TRIBUTÁRIA")
    canvas.setFont("Helvetica", 8); canvas.setFillColor(MINT)
    canvas.drawRightString(w - 18 * mm, h - 10.5 * mm, f"Cruzamento por cliente · APPA 2025 · Pág. {doc.page}")
    canvas.setStrokeColor(LINE); canvas.setLineWidth(0.5); canvas.line(18 * mm, 12 * mm, w - 18 * mm, 12 * mm)
    canvas.setFillColor(colors.HexColor("#738080")); canvas.setFont("Helvetica", 6.5)
    canvas.drawString(18 * mm, 8.5 * mm, "Uso restrito · documento gerencial. Faturamento × folha por tomador (cliente).")
    canvas.drawRightString(w - 18 * mm, 8.5 * mm, date.today().strftime("%d/%m/%Y"))
    canvas.restoreState()

_UNIF_REF = re.compile(r"ref\.?\s*\d+", re.I)
_UNIF_NUM = re.compile(r"\d{2,4}")


def carrega_uniformes(src, known):
    """Despesa de uniformes+limpeza (col PAGO) por codigo de cliente. Retorna (dict cod->direto, total_sem_empresa)."""
    if not os.path.exists(src):
        return {}, 0.0
    def _num(v):
        try:
            return float(str(v).replace(",", ".")) if v not in (None, "") else 0.0
        except Exception:
            return 0.0
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    direto = {}; sem = 0.0
    for sn in ("limpeza", "uniformes"):
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]; first = True
        for r in ws.iter_rows(values_only=True):
            if first:
                first = False; continue
            if not r or all(x is None for x in r):
                continue
            row = list(r) + [None] * 9
            v = _num(row[7])
            limpo = _UNIF_REF.sub(" ", str(row[6] or ""))
            cod = next((t for t in _UNIF_NUM.findall(limpo) if t in known), None)
            if cod:
                direto[cod] = direto.get(cod, 0.0) + v
            else:
                sem += v
    wb.close()
    return direto, sem


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data", required=True)
    ap.add_argument("--mensal", required=True)  # cruzamento_resultado.json (para o total competencia)
    ap.add_argument("--out", required=True)
    ap.add_argument("--top", type=int, default=40)
    ap.add_argument("--mes", type=int, default=0)  # 1..12; 0 = ano
    ap.add_argument("--uniformes", default=r"C:\Users\xandao\Downloads\uniformes e material de limpeza.xlsx")
    args = ap.parse_args()
    cli_json = json.load(open(args.data, encoding="utf-8"))
    rows = cli_json["rows"]
    mensal = json.load(open(args.mensal, encoding="utf-8"))["rows"]
    MESNOME = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho",
               "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    if args.mes:
        mensal = [r for r in mensal if r["competencia"] == f"2025-{args.mes:02d}"]
    # total do faturamento por competencia: usa o total do proprio cruzamento (inclui compilado 2026), com fallback
    total_fat_comp = cli_json.get("total_geral") or sum(r["faturamento_bruto"] for r in mensal)
    inss_tot = sum(r["inss_empregador"] for r in mensal)
    fgts_tot = sum(r["fgts"] for r in mensal)
    folha_tot_m = sum(r["folha_vencimentos"] for r in mensal)
    # benefícios saem da folha: encargo (INSS+FGTS) incide sobre a folha de salários (sem benefícios)
    ben_add_tot = cli_json.get("total_beneficio_add", 0.0)
    folha_salario = folha_tot_m - ben_add_tot
    enc = (inss_tot + fgts_tot) / folha_salario if folha_salario else 0.0  # encargos sobre a folha de salários

    # ---- despesas de uniformes + material de limpeza por cliente (coluna G) ----
    code2name = {str(r["client_code"]): r["cliente"] for r in rows if r.get("client_code")}
    unif_direto, unif_sem = carrega_uniformes(args.uniformes, set(code2name))
    base_rateio = sum(r["faturamento"] for r in rows) or 1.0
    for r in rows:  # despesa direta + rateio do 'sem empresa' proporcional ao faturamento
        direto = unif_direto.get(str(r.get("client_code")), 0.0)
        rateio = unif_sem * (r["faturamento"] / base_rateio)
        r["_unif"] = direto + rateio
    unif_tot = sum(r["_unif"] for r in rows)

    def situacao(fat, folha_sem, benef, unif):
        if fat <= 0:
            return None, 0.0
        custo = folha_sem + folha_sem * enc + benef + unif  # folha + encargos + benefícios + uniformes/limpeza
        marg = (fat - custo) / fat * 100
        if marg < 0:
            return "Prejuízo", marg
        if marg < 5:
            return "Margem magra", marg
        return "Lucrativo", marg

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], textColor=INK, fontSize=23, leading=26, alignment=0)
    h1s = ParagraphStyle("h1s", parent=styles["Title"], textColor=CORAL, fontSize=13, leading=15, alignment=0)
    kick = ParagraphStyle("kick", parent=styles["Normal"], textColor=CORAL, fontSize=9, fontName="Helvetica-Bold", spaceAfter=8)
    sub = ParagraphStyle("sub", parent=styles["Normal"], textColor=colors.HexColor("#36595a"), fontSize=10, leading=14)
    body = ParagraphStyle("body", parent=styles["Normal"], textColor=INK2, fontSize=8.4, leading=12.5)
    sec = ParagraphStyle("sec", parent=styles["Heading2"], textColor=INK, fontSize=13, leading=16, spaceBefore=10, spaceAfter=5)
    cli = ParagraphStyle("cli", parent=body, fontSize=6.9, leading=8.2)

    doc = BaseDocTemplate(args.out, pagesize=A4, leftMargin=15 * mm, rightMargin=15 * mm,
                          topMargin=22 * mm, bottomMargin=16 * mm)
    doc.addPageTemplates([PageTemplate(id="m", frames=[Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height)],
                                       onPage=header_footer)])
    el = [Spacer(1, 6)]
    el.append(Paragraph("RELATÓRIO GERENCIAL · CRUZAMENTO POR CLIENTE", kick))
    el.append(Paragraph("Faturamento × Folha por Tomador", h1))
    el.append(Paragraph(f"APPA · {('Competência ' + MESNOME[args.mes] + '/2025') if args.mes else 'Exercício 2025'}", h1s))
    el.append(HRFlowable(width="100%", thickness=1.6, color=CORAL, spaceBefore=6, spaceAfter=9))
    el.append(Paragraph("Relaciona, por cliente (tomador), o faturamento (regime de competência) e o custo de "
                        "mão de obra decomposto em folha, encargos, benefícios e materiais: <b>Resultado = "
                        "Faturamento − Folha − Encargos − Benefícios − Mat./Uniformes</b>.", sub))
    el.append(Spacer(1, 8))
    el.append(Paragraph(
        f"<b>Folha</b>: salários e verbas (sem benefícios). <b>Encargos</b>: INSS + FGTS ({enc*100:.1f}% sobre a folha "
        "de salários). <b>Benefícios</b>: custo líquido dos benefícios (parte da empresa − parte descontada do "
        "empregado: VT, refeição, alimentação, cesta, médico e odontológico). <b>Mat./Uniformes</b>: despesa de "
        "uniformes e material de limpeza alocada por cliente; a parcela sem cliente vinculado (interno/matriz) é "
        "rateada proporcionalmente ao faturamento. <b>Margem</b> = (faturamento − folha − encargos − benefícios − "
        "materiais) ÷ faturamento. <b>Situação</b>: "
        "<font color='#c0492f'><b>Prejuízo</b></font> (&lt;0%) · "
        "<font color='#c07a00'><b>Margem magra</b></font> (0%–5%) · "
        "<font color='#2e7d54'><b>Lucrativo</b></font> (≥5%). Faturamento por competência.", body))
    el.append(Spacer(1, 8))

    total_class = sum(r["faturamento"] for r in rows)
    total_folha = sum(r["folha_sem_beneficio"] for r in rows)
    total_benef = sum(r["beneficio_liquido"] for r in rows)
    total_unif = sum(r["_unif"] for r in rows)
    nao_id = total_fat_comp - total_class

    SIT_COLOR = {"Prejuízo": RED, "Margem magra": colors.HexColor("#c07a00"), "Lucrativo": GREEN}
    el.append(Paragraph("Clientes por faturamento", sec))
    hdr = ["#", "Cliente (tomador)", "Faturamento", "Folha", "Encargos", "Benefícios", "Mat./Unif.", "Margem", "Situação"]
    data = [hdr]
    top = rows[:args.top]
    for i, r in enumerate(top, 1):
        fol = r["folha_sem_beneficio"]; ben = r["beneficio_liquido"]; uni = r["_unif"]
        sit, marg = situacao(r["faturamento"], fol, ben, uni)
        data.append([str(i), Paragraph(r["cliente"] or r["client_code"], cli), brl(r["faturamento"]),
                     brl(fol), brl(fol * enc), brl(ben), brl(uni), f"{marg:.0f}%" if sit else "—", sit or "—"])
    demais = rows[args.top:]
    if demais:
        dfat = sum(r["faturamento"] for r in demais); dfol = sum(r["folha_sem_beneficio"] for r in demais)
        dben = sum(r["beneficio_liquido"] for r in demais); duni = sum(r["_unif"] for r in demais)
        dsit, dmarg = situacao(dfat, dfol, dben, duni)
        data.append(["", Paragraph(f"<i>Demais {len(demais)} clientes</i>", cli), brl(dfat),
                     brl(dfol), brl(dfol * enc), brl(dben), brl(duni), f"{dmarg:.0f}%" if dsit else "—", dsit or "—"])
    if nao_id > 1:
        data.append(["", Paragraph("<i>Notas sem código de cliente</i>", cli), brl(nao_id), "—", "—", "—", "—", "—", "—"])
    gsit, gmarg = situacao(total_class, total_folha, total_benef, total_unif)
    data.append(["", Paragraph("<b>TOTAL identificado</b>", cli), brl(total_class), brl(total_folha),
                 brl(total_folha * enc), brl(total_benef), brl(total_unif), f"{gmarg:.0f}%" if gsit else "—", gsit or "—"])

    cw = [5 * mm, 44 * mm, 21 * mm, 20 * mm, 18 * mm, 18 * mm, 18 * mm, 11 * mm, 16 * mm]
    t = Table(data, colWidths=cw, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), INK), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 6.3),
        ("ALIGN", (2, 0), (7, -1), "RIGHT"), ("ALIGN", (0, 0), (0, -1), "CENTER"), ("ALIGN", (8, 0), (8, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 2.2), ("BOTTOMPADDING", (0, 0), (-1, -1), 2.2),
        ("LEFTPADDING", (0, 0), (-1, -1), 2.2), ("RIGHTPADDING", (0, 0), (-1, -1), 2.2),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, PAPER]),
        ("LINEABOVE", (0, -1), (-1, -1), 0.7, INK), ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#eef3ef")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]
    for i, r in enumerate(top, 1):
        sit, marg = situacao(r["faturamento"], r["folha_sem_beneficio"], r["beneficio_liquido"], r["_unif"])
        if sit:
            c = SIT_COLOR[sit]
            style.append(("TEXTCOLOR", (7, i), (8, i), c))
            style.append(("FONTNAME", (7, i), (7, i), "Helvetica-Bold"))
    t.setStyle(TableStyle(style))
    el.append(t)
    el.append(Spacer(1, 5))
    def sitof(r):
        return situacao(r["faturamento"], r["folha_sem_beneficio"], r["beneficio_liquido"], r["_unif"])[0]
    lucr = [r for r in rows if sitof(r) == "Lucrativo"]
    prej = [r for r in rows if sitof(r) == "Prejuízo"]
    magra = [r for r in rows if sitof(r) == "Margem magra"]
    el.append(Paragraph(
        f"Faturamento 2025 (competência): <b>R$ {brl(total_fat_comp)}</b> — R$ {brl(total_class)} por cliente "
        f"({len(rows)} tomadores) + R$ {brl(nao_id)} sem código. Custo de M.O. = folha R$ {brl(total_folha)} + "
        f"encargos R$ {brl(total_folha * enc)} + benefícios R$ {brl(total_benef)} + mat./uniformes R$ {brl(total_unif)}. "
        f"Situação dos {len(rows)} tomadores: <font color='#2e7d54'><b>{len(lucr)} lucrativos</b></font>, "
        f"<font color='#c07a00'><b>{len(magra)} margem magra</b></font>, "
        f"<font color='#c0492f'><b>{len(prej)} em prejuízo</b></font>.", body))

    emis26 = cli_json.get("emitido_2026_comp_2025", 0.0)
    nnf26 = cli_json.get("notas_emitido_2026_comp_2025", 0)
    if emis26 > 0:
        nnf26_s = f"{nnf26:,}".replace(",", ".")
        el.append(Spacer(1, 4))
        el.append(Paragraph(
            f"<b>Notas faturadas em 2026 (competência 2025):</b> deste faturamento, "
            f"<b>R$ {brl(emis26)}</b> em {nnf26_s} notas foram <b>emitidas em 2026</b> mas referem-se a "
            "serviços de competência de 2025 (tomadores públicos que faturam com atraso). Pelo regime de "
            "competência, já estão somadas ao ano de 2025.", body))

    doc.build(el)
    print("OK ->", args.out)

if __name__ == "__main__":
    main()

"""Importa o relatorio "NFs Emitidas 2026.xlsx" (faturamento APPA) para o formato
mensal RETENCAO_APPA consumido pelo cruzamento Faturamento x Recebimentos.

Uso:
    python script/import_nf_billing_xlsx.py <origem.xlsx> <destino_dir> --receivables <contas_a_receber.xlsb>

O relatorio de NFs nao traz CNPJ nem codigo de cliente. O codigo e resolvido assim:
1. nome do cliente (substring, sem acentos) contra os clientes do contas a receber;
2. se ambiguo/ausente, cruza o numero da NF com os codigos candidatos;
3. sobram sem codigo -> linha e descartada e listada no relatorio final.

Gera um arquivo por mes de emissao: RETENCAO_APPA_<MM>_<MES>_<AAAA>.xlsx
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

MONTH_NAMES = {
    1: "JANEIRO",
    2: "FEVEREIRO",
    3: "MARÇO",
    4: "ABRIL",
    5: "MAIO",
    6: "JUNHO",
    7: "JULHO",
    8: "AGOSTO",
    9: "SETEMBRO",
    10: "OUTUBRO",
    11: "NOVEMBRO",
    12: "DEZEMBRO",
}

OUTPUT_HEADERS = [
    "CNPJ Cliente",
    "Cod Cliente",
    "Cliente",
    "RPS",
    "N NF E",
    "Dt Emissao",
    "Competencia",
    "Status",
    "Valor Fatura",
    "Valor INSS",
    "Valor IRRF",
    "Valor PIS",
    "Valor COFINS",
    "Valor CSLL",
    "Valor ISS",
    "Valor Liquido",
]

EXTRACTOR = Path(__file__).with_name("extract_receivables_xlsb.py")


def fail(message: str) -> None:
    raise SystemExit(f"ERRO: {message}")


def normalize(text: str) -> str:
    stripped = "".join(
        c for c in unicodedata.normalize("NFD", text.upper())
        if unicodedata.category(c) != "Mn"
    )
    return stripped.replace("º", "O").replace("ª", "A").strip()


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    fail(f"data invalida: {value!r}")


def parse_number(value):
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        fail(f"numero invalido: {value!r}")


def load_receivables(path: Path):
    result = subprocess.run(
        [sys.executable, str(EXTRACTOR), str(path)],
        capture_output=True,
        check=False,
    )
    if result.returncode != 0:
        fail(f"extractor de receivables falhou: {result.stderr.decode('utf-8', 'replace')}")
    raw = result.stdout
    if raw.startswith(b"\xff\xfe"):
        text = raw.decode("utf-16")
    else:
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = raw.decode("cp1252")
    return json.loads(text)


def build_indexes(receivables):
    name_codes: dict[str, set] = defaultdict(set)
    invoice_codes: dict[str, set] = defaultdict(set)
    for record in receivables:
        code = (record.get("client_code") or "").strip()
        if not code:
            continue
        client = (record.get("client") or "").strip()
        if client:
            name_codes[normalize(client)].add(code)
        invoice = (record.get("invoice_number") or "").strip()
        if invoice:
            invoice_codes[invoice].add(code)
    return name_codes, invoice_codes


def resolve_code(client_name: str, invoice: str | None, name_codes, invoice_codes):
    """Retorna (codigo, origem): origem = 'nome' | 'nf' | None."""
    target = normalize(client_name)
    candidates = set()
    for rich_name, codes in name_codes.items():
        if target in rich_name:
            candidates |= codes
    if len(candidates) == 1:
        return next(iter(candidates)), "nome"

    invoice = (invoice or "").strip()
    by_invoice = invoice_codes.get(invoice, set()) if invoice else set()
    if candidates:
        intersection = candidates & by_invoice
        if len(intersection) == 1:
            return next(iter(intersection)), "nf"
    elif len(by_invoice) == 1:
        return next(iter(by_invoice)), "nf"
    return None, None


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("destination", type=Path)
    parser.add_argument("--receivables", type=Path, required=True,
                        help="xlsb de contas a receber usado para resolver o codigo do cliente")
    args = parser.parse_args()

    if not args.source.exists():
        fail(f"arquivo nao encontrado: {args.source}")
    if not args.receivables.exists():
        fail(f"receivables nao encontrado: {args.receivables}")
    args.destination.mkdir(parents=True, exist_ok=True)

    print(f"lendo contas a receber: {args.receivables.name}")
    receivables = load_receivables(args.receivables)
    name_codes, invoice_codes = build_indexes(receivables)
    print(f"clientes indexados: {len(name_codes)} | notas indexadas: {len(invoice_codes)}")

    workbook = openpyxl.load_workbook(args.source, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = [str(cell).strip().lower() if cell is not None else "" for cell in next(rows)]
    if header[:3] != ["cliente", "dt emissao", "rps"]:
        fail(f"cabecalho inesperado: {header}")

    by_month: dict[tuple[int, int], list[list]] = defaultdict(list)
    stats = {"nome": 0, "nf": 0, "sem_codigo": 0, "sem_data": 0}
    unresolved: dict[str, int] = defaultdict(int)

    for row in rows:
        client = row[0]
        emission = parse_date(row[1])
        if client is None or emission is None:
            stats["sem_data"] += 1
            continue

        client_name = str(client).strip()
        invoice = str(row[3]).strip() if row[3] is not None else None
        code, origin = resolve_code(client_name, invoice, name_codes, invoice_codes)
        if code is None:
            stats["sem_codigo"] += 1
            unresolved[client_name] += 1
            continue
        stats[origin] += 1

        record = [
            None,  # CNPJ Cliente: relatorio nao traz
            code,
            client_name,
            str(row[2]).strip() if row[2] is not None else None,
            invoice,
            emission,
            parse_date(row[4]),
            str(row[5]).strip() if row[5] is not None else None,
            parse_number(row[7]),
            parse_number(row[8]),
            parse_number(row[9]),
            parse_number(row[10]),
            parse_number(row[11]),
            parse_number(row[12]),
            parse_number(row[13]),
            parse_number(row[14]),
        ]
        by_month[(emission.year, emission.month)].append(record)

    if not by_month:
        fail("nenhuma linha valida encontrada")

    for (year, month), records in sorted(by_month.items()):
        month_name = MONTH_NAMES[month]
        filename = f"RETENCAO_APPA_{month:02d}_{month_name}_{year}.xlsx"
        target = args.destination / filename

        out = openpyxl.Workbook()
        ws = out.active
        ws.append(OUTPUT_HEADERS)
        for record in records:
            ws.append(record)
        for cells in ws.iter_rows(min_row=2):
            cells[5].number_format = "DD/MM/YYYY"
            if cells[6].value is not None:
                cells[6].number_format = "DD/MM/YYYY"
        out.save(target)
        print(f"OK {filename}: {len(records)} notas")

    print()
    print(f"resolvidas por nome: {stats['nome']} | por numero da NF: {stats['nf']}")
    print(f"descartadas sem codigo: {stats['sem_codigo']} | sem cliente/data: {stats['sem_data']}")
    if unresolved:
        print("clientes sem codigo (revisar):")
        for name, count in sorted(unresolved.items(), key=lambda item: -item[1]):
            print(f"  {count:>4} notas | {name}")


if __name__ == "__main__":
    main()
